# reporter.py
import csv
from pathlib import Path
from datetime import datetime
import threading
from collections import defaultdict
from openpyxl import Workbook
from config import CSV_OUTPUT, EXCEL_OUTPUT, ADDRESSES_FILE
import os

lock = threading.Lock()
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

RESULTS_DIR = Path("results")
CSV_PATH = RESULTS_DIR / "csv"
XLSX_PATH = RESULTS_DIR / "xlsx"
CSV_PATH.mkdir(parents=True, exist_ok=True)
XLSX_PATH.mkdir(parents=True, exist_ok=True)

RAW_CSV = CSV_PATH / f"raw_data_{timestamp}.csv"
PORTFOLIO_CSV = CSV_PATH / f"portfolio_{timestamp}.csv"
XLSX_FILE = XLSX_PATH / f"portfolio_{timestamp}.xlsx"

# csv raw_data header
RAW_HEADER = ["Wallet Address", "Type", "Name", "USD Value"]

# w pamięci przechowujemy surowe wiersze, a potem agregujemy przy finalize
_raw_rows = []
_address_order = {}  # Zostanie zainicjowane przez init_reporter


def load_address_order(path: str):
    p = Path(path)
    order = {}
    if not p.exists():
        print(f"Ostrzeżenie: Plik {path} nie istnieje. Sortowanie rezerwowe.")
        return order
    with p.open("r", encoding="utf-8") as f:
        for i, line in enumerate(f):
            addr = line.strip()
            if addr:
                # Value is the index/order
                order[addr] = i
    return order


def init_reporter(address_order_map: dict):
    """Inicjuje reporter z załadowaną mapą kolejności adresów."""
    global _address_order
    _address_order = address_order_map
    print(f"Reporter zainicjowany z {len(_address_order)} adresami.")


# funkcja do zapisu pojedynczego surowego wiersza (thread-safe)
def write_raw_row(wallet_address: str, row_type: str, name: str, usd_value):
    with lock:
        # normalizacja wartości USD na float (jeśli możliwe)
        try:
            usd_f = float(usd_value)
        except Exception:
            usd_f = 0.0

        row = {
            "Wallet Address": str(wallet_address),
            "Type": str(row_type),
            "Name": str(name),
            "USD Value": usd_f,
        }
        _raw_rows.append(row)

        # append to raw csv immediately
        write_header = not RAW_CSV.exists()
        with RAW_CSV.open("a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if write_header:
                writer.writerow(RAW_HEADER)
            writer.writerow(
                [row["Wallet Address"], row["Type"], row["Name"], row["USD Value"]]
            )


# --- (Reszta funkcji jest synchroniczna i prywatna) ---


def _sort_raw_rows(rows):
    if not _address_order:
        # fallback: sortuj po adresie
        return sorted(rows, key=lambda r: r["Wallet Address"])

    def keyfn(r):
        addr = r["Wallet Address"]
        return (_address_order.get(addr, 10**9), addr)

    return sorted(rows, key=keyfn)


def _aggregate(rows):
    chain_map = defaultdict(lambda: defaultdict(float))
    portfolio_map = defaultdict(lambda: defaultdict(float))
    token_data = defaultdict(lambda: defaultdict(lambda: {"USD": 0.0, "AMOUNT": 0.0}))
    total_map = defaultdict(float)

    # 1. Process all raw rows to build the maps
    for r in rows:
        addr = r["Wallet Address"]
        typ = r["Type"].strip().lower()
        name = r["Name"]
        usd = float(r["USD Value"] or 0.0)

        if typ == "total":
            total_map[addr] = usd  # Saldo całkowite jest zbierane
        elif typ == "chain":
            chain_map[addr][name] += usd
        elif typ == "project":
            portfolio_map[addr][name] += usd
        elif typ == "token":
            parts = name.split("|")
            token_name = parts[0].strip()
            amount = 0.0
            if len(parts) >= 2:
                try:
                    amount = float(parts[1])
                except Exception:
                    pass

            token_data[token_name][addr]["USD"] += usd
            token_data[token_name][addr]["AMOUNT"] += amount

    # 2. Guarantee all addresses are present in the aggregation maps
    all_addresses = _address_order.keys()

    for addr in all_addresses:
        # Gwarantujemy włączenie w Total map (0 jeśli nie znaleziono)
        if addr not in total_map:
            total_map[addr] = 0.0

        if addr not in chain_map:
            chain_map[addr] = {}
        if addr not in portfolio_map:
            portfolio_map[addr] = {}

    return total_map, chain_map, portfolio_map, token_data


def _collect_columns(map_by_addr):
    cols = set()
    for addr, values in map_by_addr.items():
        for k in values.keys():
            cols.add(k)
    return sorted(cols)


def _write_portfolio_csv(total_map, chain_map, portfolio_map, token_map):
    chain_cols = _collect_columns(chain_map)
    port_cols = _collect_columns(portfolio_map)
    token_cols = sorted(token_map.keys())

    def sort_key(addr):
        return (_address_order.get(addr, 10**9), addr)

    all_relevant_addrs = sorted(
        set(chain_map.keys()) | set(portfolio_map.keys()) | set(_address_order.keys()),
        key=sort_key,
    )

    with PORTFOLIO_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        # 1. TOTAL block (ZAWSZE NA POCZĄTKU)
        if total_map:
            writer.writerow(["Total Balance"])
            writer.writerow(["Wallet Address", "Total USD Value"])
            for addr in all_relevant_addrs:
                # Korzystamy z get(addr, 0.0) aby obsłużyć adresy z listy bez salda (z _address_order)
                row = [addr, total_map.get(addr, 0.0)]
                writer.writerow(row)
            writer.writerow([])

        # 2. CHAIN block (Conditional)
        if chain_cols:
            writer.writerow(["Chain"])
            writer.writerow(["Wallet Address"] + chain_cols)
            for addr in all_relevant_addrs:
                row = [addr] + [chain_map.get(addr, {}).get(c, 0.0) for c in chain_cols]
                writer.writerow(row)
            writer.writerow([])

        # 3. PROJECT block (Conditional)
        if port_cols:
            writer.writerow(["Project"])
            writer.writerow(["Wallet Address"] + port_cols)
            for addr in all_relevant_addrs:
                row = [addr] + [
                    portfolio_map.get(addr, {}).get(c, 0.0) for c in port_cols
                ]
                writer.writerow(row)
            writer.writerow([])

        # 4. TOKEN block (Conditional)
        if token_cols:
            writer.writerow(["Token"])
            header = ["Wallet Address"]
            for t in token_cols:
                header.append(f"{t}_amount")
                header.append(f"{t}_USD_value")
            writer.writerow(header)

            for addr in all_relevant_addrs:
                row = [addr]
                for t in token_cols:
                    entry = token_map.get(t, {}).get(addr, {"AMOUNT": 0.0, "USD": 0.0})
                    row.append(entry["AMOUNT"])
                    row.append(entry["USD"])
                writer.writerow(row)
            writer.writerow([])


def _write_xlsx(total_map, chain_map, portfolio_map, token_map):
    # Sprawdź, czy w ogóle mamy jakieś dane, by uniknąć zapisu pustego pliku
    has_total = bool(total_map and any(total_map.values()))
    chain_cols = _collect_columns(chain_map)
    port_cols = _collect_columns(portfolio_map)
    token_cols = sorted(token_map.keys())

    if not (has_total or chain_cols or port_cols or token_cols):
        # Brak danych do zapisania, kończymy
        return

    wb = Workbook()

    def sort_key(addr):
        return (_address_order.get(addr, 10**9), addr)

    all_relevant_addrs = sorted(
        set(chain_map.keys()) | set(portfolio_map.keys()) | set(_address_order.keys()),
        key=sort_key,
    )

    # Usuń domyślny arkusz
    if wb.sheetnames == ["Sheet"]:
        wb.remove(wb.active)

    # 1. TOTAL Sheet (PIERWSZY ARKUSZ)
    if has_total:
        # Utwórz jako pierwszy arkusz (index 0)
        ws_total = wb.create_sheet("Total", 0)
        ws_total.append(["Wallet Address", "Total USD Value"])
        for addr in all_relevant_addrs:
            row = [addr, total_map.get(addr, 0.0)]
            ws_total.append(row)

    # 2. Chain sheet (Conditional)
    if chain_cols:
        ws_chain = wb.create_sheet("Chain")
        ws_chain.append(["Wallet Address"] + chain_cols)
        for addr in all_relevant_addrs:
            row = [addr] + [chain_map.get(addr, {}).get(c, 0.0) for c in chain_cols]
            ws_chain.append(row)

    # 3. Project/Portfolio sheet (Conditional)
    if port_cols:
        ws_port = wb.create_sheet("Project")
        ws_port.append(["Wallet Address"] + port_cols)
        for addr in all_relevant_addrs:
            row = [addr] + [portfolio_map.get(addr, {}).get(c, 0.0) for c in port_cols]
            ws_port.append(row)

    # 4. Token sheet (Conditional)
    if token_cols:
        ws_token = wb.create_sheet("Token")
        token_header = ["Wallet Address"]
        for t in token_cols:
            token_header.append(f"{t}_amount")
            token_header.append(f"{t}_USD_value")
        ws_token.append(token_header)

        for addr in all_relevant_addrs:
            row = [addr]
            for t in token_cols:
                entry = token_map.get(t, {}).get(addr, {"AMOUNT": 0.0, "USD": 0.0})
                row.append(entry["AMOUNT"])
                row.append(entry["USD"])
            ws_token.append(row)

    # Zapisz plik
    if wb.sheetnames:
        wb.save(XLSX_FILE)
    else:
        # Ten warunek jest dodatkowy, bo sprawdzamy na początku
        if XLSX_FILE.exists():
            os.remove(XLSX_FILE)


def finalize_outputs():
    with lock:
        if not _raw_rows and not _address_order:
            print("Brak danych do sfinalizowania.")
            return

        if not CSV_OUTPUT and not EXCEL_OUTPUT:
            raise RuntimeError(
                "Co najmniej jedno z CSV_OUTPUT lub EXCEL_OUTPUT musi być True"
            )

        print("Sortowanie surowych danych...")
        sorted_rows = _sort_raw_rows(_raw_rows)

        print("Nadpisywanie posortowanego RAW CSV...")
        with RAW_CSV.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(RAW_HEADER)
            for r in sorted_rows:
                writer.writerow(
                    [r["Wallet Address"], r["Type"], r["Name"], r["USD Value"]]
                )

        print("Agregowanie danych...")
        # Zaktualizowana nazwa i kolejność zwracanych wartości
        total_map, chain_map, portfolio_map, token_map = _aggregate(sorted_rows)

        # zapisz portfolio CSV
        if CSV_OUTPUT:
            print("Zapisywanie portfolio CSV...")
            # Dodano total_map do argumentów
            _write_portfolio_csv(total_map, chain_map, portfolio_map, token_map)

        # zapisz XLSX
        if EXCEL_OUTPUT:
            print("Zapisywanie portfolio XLSX...")
            # Dodano total_map do argumentów
            _write_xlsx(total_map, chain_map, portfolio_map, token_map)

        print("Finalizacja zakończona.")
        print(f"Surowe dane: {RAW_CSV}")
        if CSV_OUTPUT:
            print(f"Portfolio CSV: {PORTFOLIO_CSV}")
        if EXCEL_OUTPUT:
            print(f"Portfolio XLSX: {XLSX_FILE}")
